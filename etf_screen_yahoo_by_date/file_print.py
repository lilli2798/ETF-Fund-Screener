import os
import re
import pandas as pd
import openpyxl as pxl

def out_put_dir(new_file):
    try:
        os.makedirs(os.path.dirname(new_file))
    except OSError as error:
        print(error)

def update_result_book_with_create_new_with_style(fund_data, result_file, my_sheet):
    try:
        if os.path.exists(result_file):
            excel_writer = pd.ExcelWriter(result_file, engine='openpyxl', mode='a')
            fund_data.to_excel(excel_writer, index=True, sheet_name=my_sheet)
            excel_writer.close()
            # Apply styles after writing
            _apply_styles_to_sheet(result_file, my_sheet, fund_data)
        else:
            out_put_dir(result_file)
            excel_writer = pd.ExcelWriter(result_file, engine='openpyxl')
            fund_data.to_excel(excel_writer, index=True, sheet_name=my_sheet)
            excel_writer.close()
            # Apply styles after writing
            _apply_styles_to_sheet(result_file, my_sheet, fund_data)
    except Exception as e:
        print(
            'Exception in update_result_book_with_create_new_with_style for ' + result_file + ' and sheet name is ' + my_sheet)
        print(e)


def _apply_styles_to_sheet(result_file, sheet_name, fund_data):
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = pxl.load_workbook(result_file)
    ws = wb[sheet_name]

    # Set zoom level to 150%
    ws.sheet_view.zoomScale = 150

    # Freeze panes at B2 (freeze row 1 and column A)
    ws.freeze_panes = 'B2'

    # Apply word wrap to header row (row 1)
    for col in range(1, ws.max_column + 1):
        header_cell = ws.cell(row=1, column=col)
        header_cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Apply gray font to negative numbers
    gray_font = Font(color='808080')
    for row in ws.iter_rows(min_row=2):  # Skip header row
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.value < 0:
                cell.font = gray_font

    wb.save(result_file)
