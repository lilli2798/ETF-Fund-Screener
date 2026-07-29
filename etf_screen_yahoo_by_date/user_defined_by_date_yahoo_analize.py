"""
Yahoo Finance Historical Data Downloader and Analyzer (User-Defined Date Range)

This script downloads historical price data from Yahoo Finance for a list of tickers,
calculates returns over various time periods (yearly, monthly, weekly), and generates
Excel reports with performance statistics.

Features:
- Rate limiting with batch processing and random delays to avoid Yahoo API limits
- Configurable benchmark indexes for comparison (^IXIC, ^DJI, ^GSPC)
- Error handling for individual ticker failures
- Run metadata reporting (similar to ETF screener's run_recorder.yaml)
- Default file locations for automated runs
- Interactive input with fallback to defaults

Input:
- Excel file (.xlsx) with tickers as index
- Text file (.txt) with one ticker per line
- Or uses default file from config if no input provided

Output:
- Excel file with period and total return sheets
- Includes performance stats vs benchmarks
- Percentile bucket counts
- Run metadata report

Usage:
    python user_defined_by_date_yahoo_analize.py

Note: For yearly history cache management, use yearly_cache_manager.py instead.
"""

import pandas as pd
import numpy as np
import yfinance as yf
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import json
import random
import time
import yaml
from datetime import datetime
from typing import Dict, List, Optional, Tuple

from config import DEFAULT_CONFIG, DEFAULT_INDEXES
from yearly_cache_manager import to_date, make_return_series, download_adj_close


# =========================================================================
# UTILITY FUNCTIONS
# =========================================================================

def period_delta(choice):
    """Convert period choice string to pandas DateOffset.
    
    Args:
        choice: String - 'year', 'month', or 'week'
    
    Returns:
        Tuple of (DateOffset, label_string)
        - DateOffset: pandas DateOffset object for the period
        - label_string: Human-readable period name
    
    Raises:
        ValueError if choice is not year, month, or week
    """
    c = choice.lower().strip()
    if c == "year":
        return pd.DateOffset(years=1), "year"
    if c == "month":
        return pd.DateOffset(months=1), "month"
    if c == "week":
        return pd.DateOffset(weeks=1), "week"
    raise ValueError("current_choice must be year, month, or week")


def safe_tickers(vals):
    """Clean and normalize ticker symbols.
    
    Args:
        vals: Iterable of ticker symbols (strings, indices, etc.)
    
    Returns:
        List of cleaned ticker symbols (uppercase, stripped, non-NAN, non-empty)
    """
    vals = pd.Index(vals).astype(str).tolist()
    vals = [x.strip().upper() for x in vals]
    return [x for x in vals if x and x != "NAN"]


def load_input_data(input_file, config: Dict = None):
    """Load input data from Excel or text file with default fallback.
    
    Args:
        input_file: Path to input file (.xlsx or .txt), or None to use default
        config: Configuration dictionary containing default paths
    
    Returns:
        Tuple of (df, tickers, output_file, write_in_place, selected_sheet)
        - df: DataFrame with input data (index contains tickers)
        - tickers: List of cleaned ticker symbols
        - output_file: Path for output Excel file
        - write_in_place: Boolean - True if writing to input file, False if creating new
        - selected_sheet: Sheet name used (for Excel) or None (for text file)
    """
    if config is None:
        config = DEFAULT_CONFIG
    
    # If no input provided, use default
    if not input_file:
        paths_config = config.get("paths", {})
        default_input_dir = paths_config.get("default_input_dir", "../etf_screener/data/structural-data")
        
        # Look for Excel files in default directory
        default_dir = Path(default_input_dir)
        xlsx_files = list(default_dir.glob("*.xlsx"))
        
        if xlsx_files:
            # Use the most recent file
            input_file = str(max(xlsx_files, key=lambda f: f.stat().st_mtime))
            print(f"Using default input file: {input_file}")
        else:
            raise FileNotFoundError(f"No Excel files found in {default_input_dir}")
    
    input_path = Path(input_file)
    
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_file}")
    
    # Load based on file type
    if input_path.suffix.lower() == ".xlsx":
        xls = pd.ExcelFile(input_file)
        
        # Ask user which sheet to use if multiple
        if len(xls.sheet_names) > 1:
            print(f"Available sheets: {', '.join(xls.sheet_names)}")
            sheet_name = input(f"Select sheet (or press Enter for '{xls.sheet_names[0]}'): ").strip()
            if not sheet_name:
                sheet_name = xls.sheet_names[0]
        else:
            sheet_name = xls.sheet_names[0]
        
        df = pd.read_excel(input_file, sheet_name=sheet_name, index_col=0)
        df.index = df.index.astype(str)
        tickers = safe_tickers(df.index)
        
        output_file = str(input_path)
        write_in_place = True
        selected_sheet = sheet_name
        return df, tickers, output_file, write_in_place, selected_sheet
    
    elif input_path.suffix.lower() == ".txt":
        with open(input_file, "r", encoding="utf-8") as f:
            lines = [line.strip() for line in f.readlines()]
        tickers = [x.strip().upper() for x in lines if x and x != "NAN"]
        
        df = pd.DataFrame(index=tickers)
        p = input_path.with_suffix("") + "-output.xlsx"
        output_file = str(p)
        write_in_place = False
        selected_sheet = None
        return df, tickers, output_file, write_in_place, selected_sheet

    else:
        raise ValueError("Only .xlsx and .txt input files are supported")


# =========================================================================
# STATISTICS FUNCTIONS
# =========================================================================

def add_better_worst_stats(out_df, return_df, indexes: List[str] = None):
    """Add statistics comparing ticker performance to benchmark indexes.
    
    Args:
        out_df: Output DataFrame to add stats to (tickers as index)
        return_df: DataFrame of returns (tickers as index, periods as columns)
        indexes: List of benchmark index symbols (uses DEFAULT_INDEXES if None)
    
    Returns:
        DataFrame with additional columns:
        - better: Count of indexes beaten by ticker in each period
        - worst: Count of indexes underperformed by ticker in each period
        - total_columns: Total number of periods
        - no_null_columns: Number of periods with non-null data
        - better_%: Percentage of periods where ticker beat indexes
        - worst_%: Percentage of periods where ticker underperformed indexes
        - better_worst_diff: Difference between better and worst counts
    """
    if indexes is None:
        indexes = DEFAULT_INDEXES
    
    idx_return_df = return_df.loc[return_df.index.intersection(indexes)].copy()
    ticker_return_df = return_df.drop(index=indexes, errors="ignore").copy()

    highest_index = idx_return_df.max(axis=0)
    lowest_index = idx_return_df.min(axis=0)

    compare_high = ticker_return_df.gt(highest_index, axis=1)
    compare_low = ticker_return_df.lt(lowest_index, axis=1)

    stats = pd.DataFrame(index=ticker_return_df.index)
    stats["better"] = compare_high.sum(axis=1)
    stats["worst"] = compare_low.sum(axis=1)
    stats["total_columns"] = len(ticker_return_df.columns)
    stats["no_null_columns"] = ticker_return_df.count(axis=1)
    stats["better_%"] = np.where(stats["no_null_columns"] > 0, stats["better"] / stats["no_null_columns"], np.nan)
    stats["worst_%"] = np.where(stats["no_null_columns"] > 0, stats["worst"] / stats["no_null_columns"], np.nan)
    stats["better_worst_diff"] = stats["better"] - stats["worst"]

    return out_df.join(stats, how="left")


def add_percentile_bucket_counts(out_df, return_df, indexes: List[str] = None):
    """Add percentile bucket counts for each ticker's performance.
    
    Args:
        out_df: Output DataFrame to add stats to (tickers as index)
        return_df: DataFrame of returns (tickers as index, periods as columns)
        indexes: List of benchmark index symbols to exclude (uses DEFAULT_INDEXES if None)
    
    Returns:
        DataFrame with additional columns showing count of periods in each percentile bucket:
        - top_10%, top_20%, top_30%, top_40%, top_50%, top_60%, top_70%, top_80%, top_90%
        - worse_10%
        Percentiles are calculated across all tickers for each period.
    """
    if indexes is None:
        indexes = DEFAULT_INDEXES
    
    ticker_return_df = return_df.drop(index=indexes, errors="ignore").copy()
    pct_rank = ticker_return_df.rank(axis=0, method="average", pct=True)

    stats = pd.DataFrame(index=ticker_return_df.index)
    stats["top_10%"] = ((pct_rank > 0.90) & (pct_rank <= 1.00)).sum(axis=1)
    stats["top_20%"] = ((pct_rank > 0.80) & (pct_rank <= 0.90)).sum(axis=1)
    stats["top_30%"] = ((pct_rank > 0.70) & (pct_rank <= 0.80)).sum(axis=1)
    stats["top_40%"] = ((pct_rank > 0.60) & (pct_rank <= 0.70)).sum(axis=1)
    stats["top_50%"] = ((pct_rank > 0.50) & (pct_rank <= 0.60)).sum(axis=1)
    stats["top_60%"] = ((pct_rank > 0.40) & (pct_rank <= 0.50)).sum(axis=1)
    stats["top_70%"] = ((pct_rank > 0.30) & (pct_rank <= 0.40)).sum(axis=1)
    stats["top_80%"] = ((pct_rank > 0.20) & (pct_rank <= 0.30)).sum(axis=1)
    stats["top_90%"] = ((pct_rank > 0.10) & (pct_rank <= 0.20)).sum(axis=1)
    stats["worse_10%"] = (pct_rank <= 0.10).sum(axis=1)

    return out_df.join(stats, how="left")


def append_index_validation_rows(out_df, return_df, indexes: List[str] = None):
    """Append benchmark index rows for validation/comparison.
    
    Args:
        out_df: Output DataFrame to append to
        return_df: DataFrame of returns containing index data
        indexes: List of benchmark index symbols to append (uses DEFAULT_INDEXES if None)
    
    Returns:
        DataFrame with index rows appended at the bottom, separated by a blank row.
        Index rows have NaN for columns not present in their original data.
    """
    if indexes is None:
        indexes = DEFAULT_INDEXES
    
    idx_rows = return_df.loc[return_df.index.intersection(indexes)].copy()

    if idx_rows.empty:
        return out_df

    for col in out_df.columns:
        if col not in idx_rows.columns:
            idx_rows[col] = np.nan

    idx_rows = idx_rows[out_df.columns]
    blank = pd.DataFrame([[np.nan] * len(out_df.columns)], columns=out_df.columns, index=[""])
    final_df = pd.concat([out_df, blank, idx_rows], axis=0)
    return final_df


# =========================================================================
# OVERVIEW BUILDING
# =========================================================================

def build_overviews(df, adj, beginning_date, ending_date, choice, n, indexes: List[str] = None):
    """Build period and total return overview sheets.
    
    Args:
        df: Input DataFrame with ticker metadata (tickers as index)
        adj: DataFrame with adjusted close prices (tickers as columns, dates as index)
        beginning_date: Start date for analysis
        ending_date: End date for analysis
        choice: Period type - 'year', 'month', or 'week'
        n: Number of periods to analyze
        indexes: List of benchmark index symbols (uses DEFAULT_INDEXES if None)
    
    Returns:
        Tuple of (period_sheet_name, total_sheet_name, period_df, total_df, available, missing)
        - period_sheet_name: Name for period overview sheet
        - total_sheet_name: Name for total return overview sheet
        - period_df: DataFrame with rolling period returns and stats
        - total_df: DataFrame with cumulative returns from each period start to end
        - available: List of tickers with data available
        - missing: List of tickers with no data available
    """
    if indexes is None:
        indexes = DEFAULT_INDEXES
    
    delta, label = period_delta(choice)
    tickers = safe_tickers(df.index)

    all_symbols = list(dict.fromkeys(tickers + indexes))
    available = [t for t in all_symbols if t in adj.columns]
    missing = [t for t in all_symbols if t not in adj.columns]

    print(f"  Available: {len(available)}, Missing: {len(missing)}")

    beginning_dt = to_date(beginning_date)
    ending_dt = to_date(ending_date)

    period_returns = {}
    total_returns = {}

    for i in range(n):
        period_start = ending_dt - (delta * (i + 1))
        period_end = ending_dt - (delta * i)

        if period_start < beginning_dt:
            print(f"  Warning: Period {i+1} start {period_start.date()} is before beginning_date {beginning_dt.date()}")
            break

        for ticker in available:
            if ticker not in adj.columns:
                continue

            ticker_adj = adj[[ticker]]
            period_return = make_return_series(ticker_adj, period_start, period_end)
            total_return = make_return_series(ticker_adj, period_start, ending_dt)

            period_key = f"{label}_{i+1}"
            total_key = f"total_{i+1}"

            if not period_return.empty and ticker in period_return.index:
                period_returns.setdefault(ticker, {})[period_key] = period_return[ticker]

            if not total_return.empty and ticker in total_return.index:
                total_returns.setdefault(ticker, {})[total_key] = total_return[ticker]

    period_df = pd.DataFrame.from_dict(period_returns, orient='index')
    total_df = pd.DataFrame.from_dict(total_returns, orient='index')

    # Add statistics
    if not period_df.empty:
        period_df = add_better_worst_stats(period_df, period_df, indexes)
        period_df = add_percentile_bucket_counts(period_df, period_df, indexes)
        period_df = append_index_validation_rows(period_df, period_df, indexes)

    if not total_df.empty:
        total_df = add_better_worst_stats(total_df, total_df, indexes)
        total_df = add_percentile_bucket_counts(total_df, total_df, indexes)
        total_df = append_index_validation_rows(total_df, total_df, indexes)

    return (
        f"{n}-{label}-overview",
        f"Total-{n}-{label}-overview",
        period_df,
        total_df,
        available,
        missing
    )


# =========================================================================
# EXCEL OUTPUT
# =========================================================================

def add_sheet_as_table(ws, df):
    """Add DataFrame to Excel worksheet as a formatted table.
    
    Args:
        ws: Excel worksheet object
        df: DataFrame to write to the worksheet
    
    Returns:
        None (modifies worksheet in place)
    
    Features:
        - Creates Excel table with TableStyleMedium2 style
        - Freezes header row
        - Auto-adjusts column widths
        - Formats percentage columns appropriately
    """
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=1, column=j, value=col)

    for i, row in enumerate(df.itertuples(index=False), start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=None if pd.isna(val) else val)

    max_row = ws.max_row
    max_col = ws.max_column
    ref = f"A1:{get_column_letter(max_col)}{max_row}"

    tab = Table(displayName=f"Table_{ws.title}", ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)

    ws.freeze_panes = "A2"

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Format percentage columns
    pct_cols = [c for c in df.columns if "%" in str(c)]
    return_cols = [c for c in df.columns if any(k in str(c).lower() for k in ["return", "_0_", "_1_", "_2_", "_3_", "_4_", "_5_", "_6_", "_7_", "_8_", "_9_"])]

    for c in range(1, max_col + 1):
        col_name = ws.cell(row=1, column=c).value

        if col_name in pct_cols:
            for r in range(2, max_row + 1):
                ws.cell(r, c).number_format = "0.0%"
        elif col_name in return_cols:
            for r in range(2, max_row + 1):
                ws.cell(r, c).number_format = "0.00%"


def open_or_create_workbook(output_file, write_in_place):
    """Open existing workbook or create new one.
    
    Args:
        output_file: Path to Excel file
        write_in_place: If True, open existing file; if False, create new
    
    Returns:
        openpyxl Workbook object
    """
    if write_in_place and Path(output_file).exists():
        return load_workbook(output_file)

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    return wb


def save_workbook_with_retry(wb, file_path):
    """Save workbook with retry logic for file access errors.
    
    Args:
        wb: openpyxl Workbook object to save
        file_path: Path to save the workbook
    
    Returns:
        None (saves file and prompts user to retry if file is locked)
    
    Features:
        - Handles PermissionError when file is open in another application
        - Prompts user to close file and retry
    """
    while True:
        try:
            wb.save(file_path)
            return
        except PermissionError:
            print(f"\nThe file '{file_path}' is open or locked.")
            ans = input("Please close the file, then type Y to retry or 99 to stop: ").strip().lower()
            if ans == "99":
                raise SystemExit("Stopped by user because the workbook is open.")
            if ans != "y":
                print("Type Y after closing the file, or 99 to stop.")


def write_result_to_excel(output_file, write_in_place, period_sheet, total_sheet, period_out, total_out):
    """Write period and total return sheets to Excel file.
    
    Args:
        output_file: Path to Excel file
        write_in_place: If True, write to existing file; if False, create new
        period_sheet: Name for period overview sheet
        total_sheet: Name for total return overview sheet
        period_out: DataFrame with period overview data
        total_out: DataFrame with total return data
    
    Returns:
        None (writes to Excel file)
    
    Features:
        - Replaces existing sheets with same names
        - Handles file access errors with retry logic
    """
    while True:
        try:
            wb = open_or_create_workbook(output_file, write_in_place)

            if period_sheet in wb.sheetnames:
                del wb[period_sheet]
            ws1 = wb.create_sheet(period_sheet)
            add_sheet_as_table(ws1, period_out.reset_index().rename(columns={"index": "Ticker"}))

            if total_sheet in wb.sheetnames:
                del wb[total_sheet]
            ws2 = wb.create_sheet(total_sheet)
            add_sheet_as_table(ws2, total_out.reset_index().rename(columns={"index": "Ticker"}))

            save_workbook_with_retry(wb, output_file)
            wb.close()
            return

        except PermissionError:
            print(f"\nThe file '{output_file}' is open or locked.")
            ans = input("Please close the file, then type Y to retry or 99 to stop: ").strip().lower()
            if ans == "99":
                raise SystemExit("Stopped by user because the workbook is open.")
        except Exception as e:
            print(f"\nError while writing sheets: {e}")
            ans = input("Type Y to retry this write, S to skip, or 99 to stop: ").strip().lower()
            if ans == "99":
                raise SystemExit("Stopped by user.")
            if ans == "s":
                return


# =========================================================================
# REPORT GENERATION
# =========================================================================

def generate_run_report(
    input_file: str,
    beginning_date: str,
    ending_date: str,
    tickers: List[str],
    choices_made: List[Dict],
    output_file: str,
    config: Dict = None
) -> None:
    """Generate a run metadata report similar to run_recorder.yaml.
    
    Args:
        input_file: Path to the input file used
        beginning_date: Start date for the analysis
        ending_date: End date for the analysis
        tickers: List of ticker symbols processed
        choices_made: List of dictionaries tracking user choices (period, n, sheets created)
        output_file: Path to the output Excel file
        config: Configuration dictionary
    
    Returns:
        None (saves report to YAML file)
    """
    if config is None:
        config = DEFAULT_CONFIG
    
    report_config = config.get("report", {})
    if not report_config.get("enable_report", True):
        return
    
    report_path = report_config.get("report_path", "output/run_recorder.yaml")
    
    report = {
        "timestamp": datetime.now().isoformat() if report_config.get("include_timestamp", True) else None,
        "input_file": input_file,
        "query_parameters": {
            "beginning_date": beginning_date if report_config.get("include_query_params", True) else None,
            "ending_date": ending_date if report_config.get("include_query_params", True) else None,
            "num_tickers": len(tickers),
        },
        "choices_made": choices_made,
        "output_file": output_file,
    }

    report_dir = Path(report_path).parent
    report_dir.mkdir(parents=True, exist_ok=True)
    
    with open(report_path, "w", encoding="utf-8") as f:
        import yaml
        yaml.dump(report, f, default_flow_style=False)
    
    print(f"  Run report saved to {report_path}")


# =========================================================================
# MAIN ENTRY POINT
# =========================================================================

def main():
    """Main entry point for interactive Yahoo Finance data analysis.
    
    Features:
        - Interactive mode: Download data and analyze returns over custom periods
        - File access error handling with retry logic
        - Run metadata report generation
    
    User Flow:
        1. Provide input file, dates
        2. Download Yahoo Finance data
        3. Choose analysis periods (year/month/week) and number of periods
        4. Generate Excel reports with performance statistics
        5. Save run metadata report
    
    Returns:
        None
    
    Note: For yearly history cache management, use yearly_cache_manager.py instead.
    """
    config = DEFAULT_CONFIG
    choices_made = []
    
    # Interactive mode
    try:
        query_config = config.get("query", {})
        paths_config = config.get("paths", {})
        
        # Get input file with validation
        while True:
            input_file = input("input file (.xlsx or .txt, or press Enter for default): ").strip()
            if input_file:
                break
            # Use default if user pressed Enter
            default_input_dir = paths_config.get("default_input_dir", "../etf_screener/data/structural-data")
            default_dir = Path(default_input_dir)
            xlsx_files = list(default_dir.glob("*.xlsx"))
            if xlsx_files:
                print(f"Will use default input file from {default_input_dir}")
                break
            else:
                print(f"No default files found in {default_input_dir}. Please provide a file path.")
        
        # Get beginning_date with validation
        while True:
            beginning_date = input("beginning_date (yyyy-mm-dd, or press Enter for default): ").strip()
            if beginning_date:
                try:
                    pd.to_datetime(beginning_date)
                    break
                except ValueError:
                    print("Invalid date format. Please use yyyy-mm-dd format.")
            else:
                beginning_date = query_config.get("beginning_date", "2010-01-01")
                print(f"Using default beginning_date: {beginning_date}")
                break
        
        # Get ending_date with validation
        while True:
            ending_date = input("ending_date (yyyy-mm-dd, or press Enter for today): ").strip()
            if ending_date:
                try:
                    pd.to_datetime(ending_date)
                    break
                except ValueError:
                    print("Invalid date format. Please use yyyy-mm-dd format.")
            else:
                ending_date = query_config.get("ending_date")
                if ending_date is None:
                    ending_date = datetime.now().strftime("%Y-%m-%d")
                print(f"Using ending_date: {ending_date}")
                break

        # Load input data with retry for file access errors
        max_retries = 3
        for attempt in range(max_retries):
            try:
                df, tickers, output_file, write_in_place, selected_sheet = load_input_data(input_file, config)
                break
            except (FileNotFoundError, PermissionError, IOError) as e:
                if attempt < max_retries - 1:
                    print(f"\nError accessing file: {e}")
                    print("This may be because the file is open in another application.")
                    response = input("Close the file and press Enter to retry, or 'q' to quit: ").strip().lower()
                    if response == 'q':
                        return
                else:
                    print(f"\nFailed to access file after {max_retries} attempts: {e}")
                    return

        print(f"Loaded {len(tickers)} tickers")
        if selected_sheet is not None:
            print(f"Using sheet: {selected_sheet}")

        print("Downloading Yahoo data once...")
        adj = download_adj_close(tickers, beginning_date, ending_date, config)

        if adj.empty:
            print("No Yahoo data downloaded. Stop.")
            return

        while True:
            current_choice = input("\ncurrent_choice (year / month / week, or 99 to stop): ").strip().lower()
            if current_choice == "99":
                print("Done.")
                break

            if current_choice not in {"year", "month", "week"}:
                print("Invalid choice. Please type 'year', 'month', 'week', or '99' to stop.")
                continue

            while True:
                n_text = input("num_of_year / num_of_month / num_of_weeks: ").strip()
                if n_text == "99":
                    print("Done.")
                    break

                try:
                    n = int(n_text)
                    if n <= 0:
                        print("Please enter a positive integer.")
                        continue
                    break  # Valid input, proceed
                except ValueError:
                    print("Invalid integer. Please enter a number (e.g., 3, 5, 10).")
            
            if n_text == "99":
                break

            try:
                indexes = config.get("indexes")
                period_sheet, total_sheet, period_out, total_out, available, missing = build_overviews(
                    df, adj, beginning_date, ending_date, current_choice, n, indexes
                )

                write_result_to_excel(
                    output_file,
                    write_in_place,
                    period_sheet,
                    total_sheet,
                    period_out,
                    total_out
                )

                print(f"Added/Replaced sheets: {period_sheet}, {total_sheet}")
                print(f"Output file: {output_file}")
                print(f"Available tickers: {len(available)}, Missing: {len(missing)}")

                # Track choice for report
                choices_made.append({
                    "choice": current_choice,
                    "n": n,
                    "period_sheet": period_sheet,
                    "total_sheet": total_sheet,
                    "available_count": len(available),
                    "missing_count": len(missing)
                })

            except Exception as e:
                print(f"\nError while processing choice '{current_choice}' with n={n}: {e}")
                ans = input("Type Y to continue, or 99 to stop: ").strip().lower()
                if ans == "99":
                    break

        # Generate run report
        if choices_made:
            generate_run_report(
                input_file=input_file or "default",
                beginning_date=beginning_date,
                ending_date=ending_date,
                tickers=tickers,
                choices_made=choices_made,
                output_file=output_file,
                config=config
            )

    except FileNotFoundError as e:
        print(f"File not found: {e}")
    except SystemExit as e:
        print(str(e))
    except Exception as e:
        print(f"Unexpected error: {e}")


if __name__ == "__main__":
    main()
