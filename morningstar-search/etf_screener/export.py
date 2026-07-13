"""
Export helpers: writing the ranked results to Excel, formatting column
names, applying header styling (bold + wrap), and building timestamped
output paths.
"""

import time
from typing import List, Optional
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import yaml
import os
from pathlib import Path
from typing import Any, Dict


def append_to_recorder(output_dir: str, output_filename: str, thresholds: Dict[str, Any]) -> str:
    """
    Append output_filename -> thresholds (which already includes the nested
    'weights' sub-dict) into a persistent run_recorder.yaml in output_dir.
    Creates the file/dir if missing. Never overwrites prior entries.
    """
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    recorder_path = os.path.join(output_dir, "run_recorder.yaml")

    if os.path.exists(recorder_path):
        with open(recorder_path, "r") as f:
            recorder_data = yaml.safe_load(f) or {}
    else:
        recorder_data = {}

    recorder_data[output_filename] = thresholds

    with open(recorder_path, "w") as f:
        yaml.safe_dump(recorder_data, f, default_flow_style=False, sort_keys=False)

    return recorder_path

def build_timestamped_output_path(out_path: str, prefix: str = "results") -> str:
    """
    Build a full output file path of the form:
      {out_path}/{prefix}_{timestamp}.xlsx

    Returns the FULL path (directory + filename), matching how main.py
    calls this (build_timestamped_output_path(out_path, prefix=...)).
    Use os.path.basename() on the result if you need just the filename
    (e.g. for the recorder key).
    """
    from datetime import datetime
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"{prefix}_{timestamp}.xlsx"
    return os.path.join(out_path, filename)


def write_excel_with_retry(
    df: pd.DataFrame,
    path: str,
    max_retries: int = 3,
    retry_delay_seconds: float = 2.0,
) -> None:
    """
    Write `df` to `path` as an Excel file, retrying if the file is
    currently open elsewhere and locked (common on Windows/OneDrive when
    the previous output is still open in Excel).

    On PermissionError, waits `retry_delay_seconds` and tries again, up to
    `max_retries` attempts. If all retries are exhausted, falls back to
    writing under an alternate filename (appending '_retry') so the run
    doesn't lose its results entirely -- and prints a clear warning so
    you know a fallback path was used.

    Raises the original exception if the fallback write also fails.
    """
    directory: str = os.path.dirname(path)
    if directory:
        os.makedirs(directory, exist_ok=True)

    last_error: Optional[Exception] = None
    for attempt in range(1, max_retries + 1):
        try:
            df.to_excel(path, index=False)
            print(f"  Wrote {len(df)} row(s), {len(df.columns)} column(s) to: {path}")
            return
        except PermissionError as e:
            last_error = e
            print(
                f"  Attempt {attempt}/{max_retries}: '{path}' appears to be open "
                f"elsewhere ({e}). Retrying in {retry_delay_seconds}s..."
            )
            time.sleep(retry_delay_seconds)

    # All retries exhausted -- fall back to an alternate filename rather
    # than losing the computed results.
    stem, ext = os.path.splitext(path)
    fallback_path: str = f"{stem}_retry{ext}"
    print(
        f"  Warning: could not write to '{path}' after {max_retries} attempts "
        f"(file likely locked). Writing to fallback path instead: {fallback_path}"
    )
    try:
        df.to_excel(fallback_path, index=False)
        print(f"  Wrote {len(df)} row(s), {len(df.columns)} column(s) to: {fallback_path}")
    except Exception:
        print(f"  Error: fallback write to '{fallback_path}' also failed. Raising original error.")
        raise last_error


def format_column_names_for_export(df: pd.DataFrame) -> pd.DataFrame:
    """
    Convert internal snake_case/underscore column names into readable,
    title-cased headers for the exported Excel file. Does not modify the
    original DataFrame in place -- returns a renamed copy.

    Rules applied:
      - Underscores become spaces.
      - Each word is title-cased, except for known acronyms/abbreviations
        (e.g. 'ETF', 'AUM', 'YTD') which are kept fully uppercase.
      - Original Morningstar column names (which already contain spaces
        and proper casing, e.g. "Net Expense Ratio") are left untouched.
    """
    preserve_uppercase = {
        "etf", "aum", "ytd", "qtd", "sec", "roic", "nav", "cgd",
    }

    def humanize(col: str) -> str:
        if "_" not in col:
            return col
        words = col.split("_")
        formatted_words = [
            w.upper() if w.lower() in preserve_uppercase else w.capitalize()
            for w in words
        ]
        return " ".join(formatted_words)

    renamed_df: pd.DataFrame = df.rename(columns={c: humanize(c) for c in df.columns})
    return renamed_df


def apply_header_formatting(
    path: str,
    wrap_column_width: int = 18,
    header_row_height: int = 45,
    freeze_panes: str = "B2",   # freezes column A + row 1
    zoom_percent: int = 150,
) -> None:

    """
    Open a saved Excel file and apply:
      - Bold + wrap-text formatting to the header row (row 1)
      - A reasonable fixed column width so wrapped headers actually wrap
      - AutoFilter dropdowns on every column in the header row
      - Freeze panes below the header row (default "A2" -- freezes row 1,
        so headers stay visible while scrolling)
      - A worksheet zoom level (default 150%)

    This runs AFTER pandas has already written the data with to_excel(),
    since pandas itself doesn't support cell styling.

    Gracefully no-ops (with a printed warning) if the file can't be opened
    or re-saved -- this is intentionally non-fatal, since formatting is
    cosmetic and shouldn't abort a pipeline run that already successfully
    saved the underlying data.
    """
    try:
        wb = load_workbook(path)
    except Exception as e:  # noqa: BLE001
        print(f"Warning: could not open '{path}' for header formatting: {e}")
        return

    ws = wb.active
    bold_wrap_font = Font(bold=True)
    wrap_alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    max_col: int = ws.max_column
    max_row: int = ws.max_row
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = bold_wrap_font
        cell.alignment = wrap_alignment
        col_letter: str = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = wrap_column_width

    ws.row_dimensions[1].height = header_row_height

    # AutoFilter dropdowns across the full header + data range.
    last_col_letter: str = get_column_letter(max_col)
    ws.auto_filter.ref = f"A1:{last_col_letter}{max_row}"

    # Freeze panes -- everything above/left of this cell stays fixed
    # while scrolling. "A2" freezes just row 1 (the header row).
    ws.freeze_panes = freeze_panes

    # Zoom level for the worksheet view.
    ws.sheet_view.zoomScale = zoom_percent

    try:
        wb.save(path)
        print(
            f"Applied header formatting (bold+wrap, AutoFilter, freeze panes "
            f"at {freeze_panes}, zoom {zoom_percent}%) to: {path}"
        )
    except PermissionError:
        print(f"\nPermissionError: '{path}' is open elsewhere; could not apply header formatting.")

