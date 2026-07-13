"""
Profile A ETF Screening Model
==============================
Builds a long-term, low-risk, steady-return ETF shortlist from two
deduplicated Morningstar exports:

  - python-etfs.xlsx      -> structural/fundamental data only
                              (costs, sectors, managers, tax, structure
                              flags, tracking error, Sharpe, risk score)
  - performance-etfs.xlsx -> performance/risk data only
                              (multi-horizon returns, return ranks,
                              Morningstar ratings, standard deviation,
                              capture ratios, drawdowns)

The two files no longer share overlapping metric columns (only Ticker
is common), so the merge is a straightforward one-to-one join.

Both loaders accept EITHER a single file path OR a directory path. If a
directory is given, all files matching the expected naming pattern inside
it are read and combined (deduplicated on Ticker, most recent file wins).
"""

from __future__ import annotations

import difflib
import glob
import re
from pathlib import Path
from typing import List, Optional, Tuple
import numpy as np
import pandas as pd
import socket
import time
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# 1. USER-CONFIGURABLE DEFAULTS
#    (You'll be prompted for these; press Enter to accept the default shown)
# ---------------------------------------------------------------------------
DEFAULT_STRUCT_PATH: str = "*.xlsx"       # file OR folder
DEFAULT_PERF_PATH: str = "*.xlsx"     # file OR folder
DEFAULT_OUT_FILE: str = "etfs_profile_A_screened.xlsx"
DEFAULT_TOP_N_PER_CATEGORY: int = 10

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def apply_header_formatting(path: str, wrap_column_width: int = 18, header_row_height: int = 45) -> None:
    """
    Open a saved Excel file and apply bold + wrap-text formatting to the
    header row (row 1). Also sets a reasonable fixed column width so
    wrapped headers actually wrap instead of just widening the column, and
    increases the header row height so wrapped text is fully visible.

    This runs AFTER pandas has already written the data with to_excel(),
    since pandas itself doesn't support cell styling.
    """
    try:
        wb = load_workbook(path)
    except Exception as e:  # noqa: BLE001
        print(f"Warning: could not open '{path}' for header formatting: {e}")
        return

    ws = wb.active  # the sheet pandas wrote to (first/only sheet)
    bold_wrap_font = Font(bold=True)
    wrap_alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    max_col: int = ws.max_column
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = bold_wrap_font
        cell.alignment = wrap_alignment
        col_letter: str = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = wrap_column_width

    ws.row_dimensions[1].height = header_row_height

    try:
        wb.save(path)
        print(f"Applied bold + wrap-text header formatting to: {path}")
    except PermissionError:
        print(f"\nPermissionError: '{path}' is open elsewhere; could not apply header formatting.")


# ---------------------------------------------------------------------------
# 2. FILE RESOLUTION HELPERS (typo tolerance + locked-file retry)
# ---------------------------------------------------------------------------
def suggest_similar_file(bad_path: str) -> Optional[List[str]]:
    """Look in the parent folder for filenames similar to bad_path (typo help)."""
    p: Path = Path(bad_path)
    parent: Path = p.parent if str(p.parent) != "" else Path(".")
    if not parent.exists():
        return None
    candidates: List[str] = [f.name for f in parent.iterdir() if f.is_file()]
    if not candidates:
        return None
    matches: List[str] = difflib.get_close_matches(p.name, candidates, n=3, cutoff=0.5)
    return [str(parent / m) for m in matches] if matches else None


def resolve_path_interactively(path: str) -> str:
    """
    Keep asking the user until `path` points to something that exists
    (either a file or a directory), offering typo-corrected suggestions
    found nearby when possible.
    """
    while not Path(path).exists():
        print(f"\nPath not found: {path}")
        suggestions: Optional[List[str]] = suggest_similar_file(path)
        if suggestions:
            print("Did you mean one of these?")
            for i, s in enumerate(suggestions, start=1):
                print(f"  {i}. {s}")
            choice: str = input(
                "Enter the number of the correct file, type a new path, "
                "or press Enter to cancel: "
            ).strip()
            if choice.isdigit() and 1 <= int(choice) <= len(suggestions):
                path = suggestions[int(choice) - 1]
                continue
            elif choice == "":
                raise FileNotFoundError(f"No valid path selected for: {path}")
            else:
                path = choice
                continue
        else:
            new_path: str = input(
                "No similar files found nearby. Please re-enter the correct "
                "path (or press Enter to cancel): "
            ).strip()
            if not new_path:
                raise FileNotFoundError(f"No valid path selected for: {path}")
            path = new_path
    return path


def resolve_files_in_path(path: str, file_pattern: str, exclude_dir: Optional[str] = None) -> List[str]:
    """
    Resolve `path` into a list of Excel files to load.

    - If `path` is a single existing file, return [path].
    - If `path` is a directory, glob for files matching `file_pattern`
      (default "*.xlsx" -> any Excel file) inside it.
    - Files located inside `exclude_dir` (e.g. your results/output folder,
      if nested under the input directory) are skipped so previously
      generated output files never get re-ingested as input.
    - Skips Excel "lock" files (start with '~$') that Excel creates while
      a workbook is open.
    """
    p: Path = Path(path)

    if p.is_file():
        return [str(p)]

    if p.is_dir():
        matches: List[str] = sorted(glob.glob(str(p / file_pattern)))
        # Filter out Excel lock files and anything inside the excluded output dir
        exclude_resolved: Optional[Path] = Path(exclude_dir).resolve() if exclude_dir else None
        filtered: List[str] = []
        for m in matches:
            m_path = Path(m)
            if m_path.name.startswith("~$"):
                continue
            if exclude_resolved and exclude_resolved in m_path.resolve().parents:
                continue
            filtered.append(m)

        if not filtered:
            raise FileNotFoundError(
                f"No files matching '{file_pattern}' were found in directory: {p}"
            )
        print(f"Found {len(filtered)} file(s) matching '{file_pattern}' in {p}:")
        for m in filtered:
            print(f"  - {m}")
        return filtered

    resolved: str = resolve_path_interactively(str(p))
    return resolve_files_in_path(resolved, file_pattern, exclude_dir)


import socket
import time


def read_excel_with_retry(path: str, usecols: Optional[List[str]] = None) -> pd.DataFrame:
    """
    Open a single Excel file robustly:
      - Resolves typos / missing files before attempting to read.
      - Retries automatically if the file is open/locked (PermissionError).
      - Retries with backoff on TimeoutError, which commonly means a
        cloud-storage file (OneDrive/iCloud) hasn't finished downloading
        to local disk yet.
      - Falls back to no column filtering on ValueError (bad usecols).
      - Catches any other unexpected error and lets the user retry or abort.
    """
    path = resolve_path_interactively(path)
    max_timeout_retries: int = 5
    timeout_attempt: int = 0

    while True:
        try:
            df: pd.DataFrame = pd.read_excel(path, engine="openpyxl", usecols=usecols)
            return df
        except FileNotFoundError:
            path = resolve_path_interactively(path)
        except PermissionError:
            print(f"\nPermissionError: '{path}' is likely open in Excel or another program.")
            input("Please close the file, then press Enter to retry...")
        except (TimeoutError, socket.timeout) as e:
            timeout_attempt += 1
            print(
                f"\nTimeoutError reading '{path}': {e}\n"
                f"This usually means the file is cloud-only (OneDrive/iCloud) and hasn't "
                f"finished downloading locally. In Finder, right-click the file (or its "
                f"parent folder) and choose 'Always Keep on This Device', then wait for "
                f"the download to finish."
            )
            if timeout_attempt >= max_timeout_retries:
                skip = input(
                    f"Still timing out after {timeout_attempt} attempts. Press Enter to "
                    f"retry again, or type 'skip' to skip this file: "
                ).strip().lower()
                if skip == "skip":
                    raise
            wait_seconds: int = min(5 * timeout_attempt, 30)
            print(f"Waiting {wait_seconds}s before retrying...")
            time.sleep(wait_seconds)
        except ValueError as e:
            print(f"\nColumn selection issue while reading '{path}': {e}")
            print("Retrying without column filtering to inspect available columns...")
            usecols = None
        except Exception as e:  # noqa: BLE001
            print(f"\nUnexpected error while reading '{path}': {type(e).__name__}: {e}")
            retry_choice: str = input("Type a new path to retry, or press Enter to abort: ").strip()
            if not retry_choice:
                raise
            path = retry_choice


def read_and_concat_excels(
    file_list: List[str],
    usecols: Optional[List[str]] = None,
    dedupe_on: str = "Ticker",
    min_matching_cols_ratio: float = 0.5,
) -> pd.DataFrame:
    """
    Read one or more Excel files (each with retry/typo handling) and
    concatenate them into a single DataFrame.

    - Every file is tagged with a '__source_file__' column for traceability.
    - If `usecols` is provided, warns (but doesn't skip) when a file matches
      fewer than `min_matching_cols_ratio` of the expected columns -- this
      catches cases where the wrong type of export (e.g. performance data
      mixed into the structural folder) got picked up by the glob pattern.
    - Duplicate values in `dedupe_on` across files: the LAST file wins.
    """
    frames: List[pd.DataFrame] = []
    for f in file_list:
        print(f"Reading: {f}")
        df: pd.DataFrame = read_excel_with_retry(f, usecols=usecols)
        df.columns = [str(c).strip() for c in df.columns]

        if usecols:
            matched_ratio: float = len(set(df.columns) & set(usecols)) / len(usecols)
            if matched_ratio < min_matching_cols_ratio:
                print(f"  Warning: '{f}' only matches {matched_ratio:.0%} of expected "
                      f"columns for this dataset type -- double-check this file belongs "
                      f"in this folder.")

        df["__source_file__"] = Path(f).name
        frames.append(df)

    if not frames:
        raise ValueError("No files were successfully read; nothing to concatenate.")

    combined: pd.DataFrame = pd.concat(frames, ignore_index=True, sort=False)

    if dedupe_on in combined.columns:
        dupe_count: int = int(combined.duplicated(subset=dedupe_on, keep="last").sum())
        if dupe_count > 0:
            print(f"Warning: {dupe_count} duplicate '{dedupe_on}' value(s) found across "
                  f"files; keeping the row from the most recently listed file for each.")
        combined = combined.drop_duplicates(subset=dedupe_on, keep="last")
    else:
        print(f"Warning: dedupe column '{dedupe_on}' not found; skipping de-duplication.")

    return combined



def write_excel_with_retry(df: pd.DataFrame, path: str) -> None:
    """Save a DataFrame to Excel, retrying if the destination file is open/locked."""
    while True:
        try:
            df.to_excel(path, index=False)
            print(f"Saved: {path}")
            return
        except PermissionError:
            print(f"\nPermissionError: '{path}' is likely open in Excel.")
            input("Please close the file, then press Enter to retry saving...")
        except Exception as e:  # noqa: BLE001 - catch-all so a bad output path doesn't crash silently
            print(f"\nUnexpected error while saving '{path}': {type(e).__name__}: {e}")
            new_path: str = input(
                "Enter a different output path to retry, or press Enter to abort: "
            ).strip()
            if not new_path:
                raise
            path = new_path


def get_paths_from_user() -> Tuple[str, str, str, int]:
    """Prompt for input/output paths (files OR folders) and Top-N, with sensible defaults."""
    struct_path: str = input(
        f"Structural data file or folder [{DEFAULT_STRUCT_PATH}]: "
    ).strip() or DEFAULT_STRUCT_PATH
    perf_path: str = input(
        f"Performance data file or folder [{DEFAULT_PERF_PATH}]: "
    ).strip() or DEFAULT_PERF_PATH
    out_path: str = input(
        f"Output screened file [{DEFAULT_OUT_FILE}]: "
    ).strip() or DEFAULT_OUT_FILE

    top_n_raw: str = input(
        f"Top N ETFs per category [{DEFAULT_TOP_N_PER_CATEGORY}]: "
    ).strip()
    top_n: int
    try:
        top_n = int(top_n_raw) if top_n_raw else DEFAULT_TOP_N_PER_CATEGORY
    except ValueError:
        print(f"Invalid number entered, using default {DEFAULT_TOP_N_PER_CATEGORY}.")
        top_n = DEFAULT_TOP_N_PER_CATEGORY

    return struct_path, perf_path, out_path, top_n


def build_timestamped_output_path(out_path: str, default_stem: str = "etfs_profile_A_screened") -> str:
    """
    Build a timestamped output file path.

    - If out_path is an existing directory (or ends with a path separator),
      the file is created inside it using `default_stem`.
    - If out_path looks like a specific filename, its stem is kept and a
      timestamp is inserted before the extension.
    - Timestamp format: YYYYMMDD_HHMMSS
    """
    timestamp: str = datetime.now().strftime("%Y%m%d_%H%M%S")
    p: Path = Path(out_path)

    if p.is_dir() or out_path.endswith(("/", "\\")):
        p.mkdir(parents=True, exist_ok=True)
        return str(p / f"{default_stem}_{timestamp}.xlsx")

    stem: str = p.stem or default_stem
    suffix: str = p.suffix or ".xlsx"
    p.parent.mkdir(parents=True, exist_ok=True)
    return str(p.parent / f"{stem}_{timestamp}{suffix}")


def apply_etf_only_filter(df: pd.DataFrame) -> pd.DataFrame:
    """
    Keep only true ETFs, excluding individual stocks that were included in
    the Morningstar export.

    Uses multiple signals, since a single column may be blank or unreliable
    for some rows:
      1. 'Share Class Type' == 'ETF' (Morningstar's direct classification) -
         strongest signal when present.
      2. 'Morningstar Category' is a recognized fund category (stocks
         usually have this blank, or show an equity sector/industry
         classification instead of a Morningstar fund category).
      3. Fund-level structural fields must be populated (Net Expense Ratio,
         Fund Size) - individual stocks don't have expense ratios or
         "fund size"; they'd be NaN for a stock row.

    A row is kept if it passes signal 1 when available, OR passes both
    signal 2 and signal 3 as a fallback when Share Class Type is missing
    or ambiguous.
    """
    filtered: pd.DataFrame = df.copy()
    start_count: int = len(filtered)

    has_share_class_col: bool = "Share Class Type" in filtered.columns
    has_expense_col: bool = "Net Expense Ratio" in filtered.columns
    has_fund_size_col: bool = "Fund Size" in filtered.columns

    if has_share_class_col:
        is_etf_flagged: pd.Series = (
            filtered["Share Class Type"].astype(str).str.strip().str.upper() == "ETF"
        )
    else:
        is_etf_flagged = pd.Series(False, index=filtered.index)

    # Fallback heuristic: must have BOTH an expense ratio and a fund size
    # populated -- individual stocks will be NaN on both of these.
    if has_expense_col and has_fund_size_col:
        has_fund_level_data: pd.Series = (
            filtered["Net Expense Ratio"].notna() & filtered["Fund Size"].notna()
        )
    else:
        has_fund_level_data = pd.Series(True, index=filtered.index)  # can't check -> don't exclude on this basis

    # Keep rows that pass the direct flag OR the fallback heuristic
    keep_mask: pd.Series = is_etf_flagged | (~has_share_class_col & has_fund_level_data)
    # If Share Class Type exists but is something other than ETF (e.g. blank,
    # "Common Stock", "Mutual Fund"), and it also fails the fund-level data
    # check, it's excluded -- this catches stocks even when mislabeled.
    if has_share_class_col:
        keep_mask = is_etf_flagged | (filtered["Share Class Type"].isna() & has_fund_level_data)

    excluded: pd.DataFrame = filtered[~keep_mask]
    filtered = filtered[keep_mask]

    excluded_count: int = start_count - len(filtered)
    if excluded_count > 0:
        print(f"ETF-only filter: excluded {excluded_count} non-ETF row(s).")
        if "Ticker" in excluded.columns:
            sample_tickers = excluded["Ticker"].dropna().astype(str).head(10).tolist()
            print(f"  Examples of excluded tickers: {sample_tickers}")

    print(f"ETF-only filter: {start_count} -> {len(filtered)} rows remain.")
    return filtered


# ---------------------------------------------------------------------------
# 3. COLUMN SCHEMAS
#    Matches your current deduplicated exports exactly. Only these columns
#    are read from disk, which cuts I/O time and memory on large files.
# ---------------------------------------------------------------------------
STRUCT_NEEDED_COLS: List[str] = [
    # Identifiers & grouping
    "Ticker", "Morningstar Category", "Exchange", "Exchange Country",
    "Primary Benchmark",
    # Risk-adjusted performance
    "Sharpe Ratio (1Y Monthly)", "Sharpe Ratio (3Y Monthly)",
    # Volatility & downside risk (text field: label + score in parentheses)
    "Worst Three Month Return", "Portfolio Risk Score",
    # Tracking quality
    "Tracking Error (1Y Monthly)", "Tracking Error (3Y Monthly)",
    # Costs & fees
    "Net Expense Ratio", "Adjusted Expense Ratio", "Management Fee",
    # Liquidity & size
    "Total Net Assets for Share Class", "Fund Size", "Trading Volume",
    "Shares Outstanding", "Premium/Discount", "Premium/Discount (1Y Avg)",
    "Trading Status",
    # Quality, valuation & portfolio characteristics
    "Portfolio Growth Grade", "Portfolio Financial Health Grade",
    "Portfolio Economic Moat Coverage (Wide)", "Portfolio Return on Invested Capital",
    "Portfolio Price/Earnings", "Portfolio Price/Book", "Portfolio Price/Sales",
    "Portfolio Price/Free Cash Flow", "Portfolio Price/Fair Value",
    # Sector / exposure
    "Sector Allocation (Basic Materials)", "Sector Allocation (Consumer Cyclical)",
    "Sector Allocation (Financial Services)", "Sector Allocation (Real Estate)",
    "Sector Allocation (Communication Services)", "Sector Allocation (Energy)",
    "Sector Allocation (Industrials)", "Sector Allocation (Technology)",
    "Sector Allocation (Consumer Defensive)", "Sector Allocation (Healthcare)",
    "Sector Allocation (Utilities)",
    # Manager & stewardship
    "Fund Managers", "Number of Fund Managers", "Longest Manager Tenure",
    "Management Style", "Medalist Rating (Overall)",
    # Tax & income
    "Tax Cost Ratio (1Y)", "Tax Cost Ratio (2Y)", "Potential Capital Gains Exposure",
    "SEC 30-Day Yield",
    # Structure & flags
    "Leveraged Fund", "Interval Fund", "Fund of Funds", "Investment Status",
    "Total Leverage Ratio", "Strategic Beta Group", "Inception Date",
]

PERF_NEEDED_COLS: List[str] = [
    # Identifiers (Name only lives in this file now)
    "Ticker", "Name",
    # Long-term returns
    "Total Return (3Y)", "Total Return (5Y)", "Total Return (10Y)",
    "Total Return (Since Inception)",
    # Return ranks (long-term)
    "3Y Return Rank in Category", "5Y Return Rank in Category",
    "10Y Return Rank in Category",
    # Morningstar ratings
    "Morningstar Risk Rating (Overall)", "Morningstar Risk Rating (5Y)",
    "Morningstar Risk Rating (10Y)",
    "Morningstar Return Rating (Overall)",
    "Morningstar Rating for Funds (Overall)",
    # Volatility
    "Standard Deviation (3Y Monthly)", "Standard Deviation (5Y Monthly)",
    "Standard Deviation (10Y Monthly)",
    # Capture ratios
    "Upside Capture Ratio (3Y)", "Upside Capture Ratio (5Y)",
    "Downside Capture Ratio (3Y)", "Downside Capture Ratio (5Y)",
    # Drawdowns
    "Maximum Drawdown (5Y)", "Maximum Drawdown (10Y)",
    # Structure flags present only in this file
    "Index Fund", "No Load Fund", "Maximum Deferred Load", "Maximum Front Load",
]

NUMERIC_COLS_ALL: List[str] = [
    "Sharpe Ratio (1Y Monthly)", "Sharpe Ratio (3Y Monthly)",
    "Worst Three Month Return", "Tracking Error (1Y Monthly)", "Tracking Error (3Y Monthly)",
    "Net Expense Ratio", "Adjusted Expense Ratio", "Management Fee",
    "Total Net Assets for Share Class", "Fund Size", "Trading Volume",
    "Shares Outstanding", "Premium/Discount", "Premium/Discount (1Y Avg)",
    "Portfolio Return on Invested Capital", "Portfolio Price/Earnings",
    "Portfolio Price/Book", "Portfolio Price/Sales", "Portfolio Price/Free Cash Flow",
    "Portfolio Price/Fair Value",
    "Sector Allocation (Basic Materials)", "Sector Allocation (Consumer Cyclical)",
    "Sector Allocation (Financial Services)", "Sector Allocation (Real Estate)",
    "Sector Allocation (Communication Services)", "Sector Allocation (Energy)",
    "Sector Allocation (Industrials)", "Sector Allocation (Technology)",
    "Sector Allocation (Consumer Defensive)", "Sector Allocation (Healthcare)",
    "Sector Allocation (Utilities)",
    "Number of Fund Managers", "Longest Manager Tenure",
    "Tax Cost Ratio (1Y)", "Tax Cost Ratio (2Y)", "Potential Capital Gains Exposure",
    "SEC 30-Day Yield", "Total Leverage Ratio",
    "Total Return (3Y)", "Total Return (5Y)", "Total Return (10Y)",
    "Total Return (Since Inception)",
    "3Y Return Rank in Category", "5Y Return Rank in Category", "10Y Return Rank in Category",
    "Morningstar Rating for Funds (Overall)",
    "Standard Deviation (3Y Monthly)", "Standard Deviation (5Y Monthly)",
    "Standard Deviation (10Y Monthly)",
    "Upside Capture Ratio (3Y)", "Upside Capture Ratio (5Y)",
    "Downside Capture Ratio (3Y)", "Downside Capture Ratio (5Y)",
    "Maximum Drawdown (5Y)", "Maximum Drawdown (10Y)",
    "Maximum Deferred Load", "Maximum Front Load",
]

DATE_COLS_ALL: List[str] = ["Inception Date"]

# Ordinal maps for text-based rating/grade fields (higher number = better)
GRADE_MAP: dict = {
    "A+": 12, "A": 11, "A-": 10,
    "B+": 9, "B": 8, "B-": 7,
    "C+": 6, "C": 5, "C-": 4,
    "D+": 3, "D": 2, "D-": 1, "F": 0,
}
MEDALIST_MAP: dict = {"Gold": 5, "Silver": 4, "Bronze": 3, "Neutral": 2, "Negative": 1}
MSTAR_RISK_LABEL_MAP: dict = {
    "Low": 5, "Below Average": 4, "Average": 3, "Above Average": 2, "High": 1,
}


# ---------------------------------------------------------------------------
# 4. LOAD + CLEAN
# ---------------------------------------------------------------------------
from typing import Any

def parse_portfolio_risk_score(value: Any) -> Tuple[Optional[str], float]:
    """
    'Portfolio Risk Score' in python-etfs.xlsx is a combined text field like
    'Very Aggressive (87)'. Split into a text label and a numeric score.

    Returns (label, numeric_score). numeric_score is NaN if not parseable;
    label is None if the input itself is missing.
    """
    if pd.isna(value):
        return None, np.nan
    match = re.match(r"^(.*?)\s*\\((-?\\d+(?:\\.\\d+)?)\\)\s*$", str(value).strip())
    if match:
        label: str = match.group(1).strip()
        score: float = float(match.group(2))
        return label, score
    return str(value).strip(), np.nan



def clean_dtypes(df: pd.DataFrame) -> pd.DataFrame:
    """
    Force numeric/date columns to correct dtypes; bad/unparseable values
    become NaN instead of raising, so a single malformed cell never crashes
    the whole pipeline.
    """
    numeric_cols = df.columns.intersection(NUMERIC_COLS_ALL)
    for col in numeric_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    date_cols = df.columns.intersection(DATE_COLS_ALL)
    for col in date_cols:
        df[col] = pd.to_datetime(df[col], errors="coerce")

    # Trim whitespace on remaining text/object columns for consistent grouping
    obj_cols = df.select_dtypes(include="object").columns
    for c in obj_cols:
        df[c] = df[c].astype(str).str.strip().replace({"nan": np.nan, "None": np.nan})

    return df


def load_structural_data(
    path: str = DEFAULT_STRUCT_PATH,
    file_pattern: str = DEFAULT_STRUCT_PATH,
    exclude_dir: Optional[str] = None,
) -> pd.DataFrame:
    """
    Load python-etfs.xlsx (structural/fundamental fields only).

    `path` can be:
      - a single file path, or
      - a directory containing one or more files matching `file_pattern`
        (all matches are read and combined, deduplicated on Ticker).

    `exclude_dir`, if provided, skips any files located inside that
    directory (e.g. your output/results folder) so previously generated
    output files never get re-ingested as input.

    Raises ValueError if required identifier columns are missing after load.
    """
    files: List[str] = resolve_files_in_path(path, file_pattern, exclude_dir=exclude_dir)
    df: pd.DataFrame = read_and_concat_excels(files, usecols=STRUCT_NEEDED_COLS)

    required: List[str] = ["Ticker", "Morningstar Category"]
    missing: List[str] = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Structural data is missing required columns: {missing}")

    df = clean_dtypes(df)

    if "Portfolio Risk Score" in df.columns:
        parsed = df["Portfolio Risk Score"].apply(parse_portfolio_risk_score)
        df["Risk_Label"] = parsed.apply(lambda t: t[0])
        df["Risk_Score_Numeric"] = parsed.apply(lambda t: t[1])

    return df


def load_performance_data(
    path: str = DEFAULT_PERF_PATH,
    file_pattern: str = DEFAULT_PERF_PATH,
    exclude_dir: Optional[str] = None,
) -> pd.DataFrame:
    """
    Load performance-etfs.xlsx (returns, ranks, risk, capture, drawdowns).

    `path` can be a single file or a directory containing one or more files
    matching `file_pattern` (all matches are read and combined, deduplicated
    on Ticker).

    `exclude_dir`, if provided, skips any files located inside that
    directory (e.g. your output/results folder).

    Raises ValueError if the 'Ticker' join key is missing after load.
    """
    files: List[str] = resolve_files_in_path(path, file_pattern, exclude_dir=exclude_dir)
    df: pd.DataFrame = read_and_concat_excels(files, usecols=PERF_NEEDED_COLS)

    if "Ticker" not in df.columns:
        raise ValueError("Performance data is missing the required 'Ticker' column.")

    return clean_dtypes(df)



def merge_datasets(df_struct: pd.DataFrame, df_perf: pd.DataFrame) -> pd.DataFrame:
    """
    Merge structural + performance data on Ticker, coalescing any
    overlapping metric columns so no _x/_y duplicates remain.
    """
    overlap: List[str] = [
        c for c in df_struct.columns
        if c in df_perf.columns and c != "Ticker"
    ]
    if overlap:
        print(f"Note: {len(overlap)} overlapping column(s) found between "
              f"structural and performance data: {overlap}")

    merged: pd.DataFrame = pd.merge(
        df_struct, df_perf,
        on="Ticker", how="inner",
        validate="one_to_one",
        suffixes=("_struct", "_perf"),
    )

    # Coalesce: prefer performance-side values, fall back to structural-side
    for col in overlap:
        struct_col, perf_col = f"{col}_struct", f"{col}_perf"
        if struct_col in merged.columns and perf_col in merged.columns:
            merged[col] = merged[perf_col].combine_first(merged[struct_col])
            merged.drop(columns=[struct_col, perf_col], inplace=True)

    unmatched: int = len(df_struct) - len(merged)
    if unmatched > 0:
        print(
            f"Note: {unmatched} ticker(s) from the structural data had no "
            f"matching row in the performance data and were dropped."
        )

    return merged



# ---------------------------------------------------------------------------
# 5. CONCEPT SCORES (0-100, category-relative)
# ---------------------------------------------------------------------------
def percentile_rank_by_group(
    df: pd.DataFrame,
    group_col: str,
    value_col: str,
    higher_is_better: bool = True,
) -> pd.Series:
    """0-100 percentile rank of value_col within each group_col bucket."""
    if value_col not in df.columns:
        return pd.Series(np.nan, index=df.index)
    s: pd.Series = df.groupby(group_col)[value_col].rank(pct=True, na_option="keep")
    if not higher_is_better:
        s = 1.0 - s
    return s * 100.0


def safe_mean(series_list: List[Optional[pd.Series]]) -> pd.Series:
    """Row-wise mean across a list of Series, ignoring NaNs; NaN if all missing."""
    valid_series: List[pd.Series] = [s for s in series_list if s is not None]
    if not valid_series:
        return pd.Series(dtype=float)
    stacked: pd.DataFrame = pd.concat(valid_series, axis=1)
    return stacked.mean(axis=1, skipna=True)


def build_concept_scores(df: pd.DataFrame) -> pd.DataFrame:
    """Add Profile A concept-score columns to df (all 0-100, category-relative)."""
    cat: str = "Morningstar Category"

    # --- Long-term performance: returns + ranks ---
    returns_part: pd.Series = safe_mean([
        percentile_rank_by_group(df, cat, "Total Return (3Y)", True),
        percentile_rank_by_group(df, cat, "Total Return (5Y)", True),
        percentile_rank_by_group(df, cat, "Total Return (10Y)", True),
    ])
    ranks_part: pd.Series = safe_mean([
        percentile_rank_by_group(df, cat, "3Y Return Rank in Category", False),
        percentile_rank_by_group(df, cat, "5Y Return Rank in Category", False),
        percentile_rank_by_group(df, cat, "10Y Return Rank in Category", False),
    ])
    df["Performance_LT_Score"] = safe_mean([returns_part, ranks_part])

    # --- Sharpe (risk-adjusted performance) ---
    df["Sharpe_LT_Score"] = safe_mean([
        percentile_rank_by_group(df, cat, "Sharpe Ratio (3Y Monthly)", True),
        percentile_rank_by_group(df, cat, "Sharpe Ratio (1Y Monthly)", True),
    ])

    # --- Volatility & drawdown (lower is better) ---
    df["Volatility_Score"] = safe_mean([
        percentile_rank_by_group(df, cat, "Standard Deviation (3Y Monthly)", False),
        percentile_rank_by_group(df, cat, "Standard Deviation (5Y Monthly)", False),
        percentile_rank_by_group(df, cat, "Standard Deviation (10Y Monthly)", False),
        percentile_rank_by_group(df, cat, "Worst Three Month Return", True),  # less negative better
        percentile_rank_by_group(df, cat, "Maximum Drawdown (5Y)", True),
        percentile_rank_by_group(df, cat, "Maximum Drawdown (10Y)", True),
        percentile_rank_by_group(df, cat, "Risk_Score_Numeric", False),
    ])

    # --- Tracking quality (grouped by benchmark, not category) ---
    df["Tracking_Score"] = safe_mean([
        percentile_rank_by_group(df, "Primary Benchmark", "Tracking Error (3Y Monthly)", False),
        percentile_rank_by_group(df, cat, "Net Expense Ratio", False),
    ])

    # --- Costs & fees ---
    df["Cost_Score"] = safe_mean([
        percentile_rank_by_group(df, cat, "Net Expense Ratio", False),
        percentile_rank_by_group(df, cat, "Adjusted Expense Ratio", False),
        percentile_rank_by_group(df, cat, "Management Fee", False),
    ])

    # --- Quality & valuation (letter grades mapped to numeric first) ---
    for grade_col in ["Portfolio Growth Grade", "Portfolio Financial Health Grade"]:
        if grade_col in df.columns:
            df[f"{grade_col}_num"] = df[grade_col].map(GRADE_MAP)

    q_part: pd.Series = safe_mean([
        percentile_rank_by_group(df, cat, "Portfolio Growth Grade_num", True),
        percentile_rank_by_group(df, cat, "Portfolio Financial Health Grade_num", True),
        percentile_rank_by_group(df, cat, "Portfolio Return on Invested Capital", True),
    ])
    v_part: pd.Series = safe_mean([
        percentile_rank_by_group(df, cat, "Portfolio Price/Earnings", False),
        percentile_rank_by_group(df, cat, "Portfolio Price/Book", False),
        percentile_rank_by_group(df, cat, "Portfolio Price/Free Cash Flow", False),
        percentile_rank_by_group(df, cat, "Portfolio Price/Fair Value", False),
    ])
    df["Quality_Valuation_Score"] = safe_mean([q_part, v_part])

    # --- Liquidity & size ---
    df["Liquidity_Score"] = safe_mean([
        percentile_rank_by_group(df, cat, "Total Net Assets for Share Class", True),
        percentile_rank_by_group(df, cat, "Trading Volume", True),
    ])

    # --- Manager & stewardship ---
    if "Medalist Rating (Overall)" in df.columns:
        df["Medalist_Rating_num"] = df["Medalist Rating (Overall)"].map(MEDALIST_MAP)

    df["Manager_Score"] = safe_mean([
        percentile_rank_by_group(df, cat, "Medalist_Rating_num", True),
        percentile_rank_by_group(df, cat, "Morningstar Rating for Funds (Overall)", True),
        percentile_rank_by_group(df, cat, "Longest Manager Tenure", True),
    ])

    # --- Tax efficiency ---
    df["Tax_Score"] = safe_mean([
        percentile_rank_by_group(df, cat, "Tax Cost Ratio (2Y)", False),
        percentile_rank_by_group(df, cat, "Potential Capital Gains Exposure", False),
    ])

    return df


# ---------------------------------------------------------------------------
# 6. ELIGIBILITY FILTERS (Profile A: long-term, low-risk, steady)
# ---------------------------------------------------------------------------
def apply_profile_A_filters(
    df: pd.DataFrame,
    min_aum: float = 1e9,
    min_volume: float = 50_000,
) -> pd.DataFrame:
    """Hard exclusion rules for the Profile A eligible universe."""
    eligible: pd.DataFrame = df.copy()
    start_count: int = len(eligible)

    for flag_col in ["Leveraged Fund", "Interval Fund", "Fund of Funds"]:
        if flag_col in eligible.columns:
            eligible = eligible[
                eligible[flag_col].astype(str).str.strip().str.lower() != "yes"
            ]

    if "Investment Status" in eligible.columns:
        eligible = eligible[
            eligible["Investment Status"].astype(str).str.contains("Open", na=False)
        ]

    if "Total Net Assets for Share Class" in eligible.columns:
        eligible = eligible[eligible["Total Net Assets for Share Class"] >= min_aum]

    if "Trading Volume" in eligible.columns:
        eligible = eligible[eligible["Trading Volume"] >= min_volume]

    print(f"Eligibility filter: {start_count} -> {len(eligible)} ETFs remain.")
    return eligible


# ---------------------------------------------------------------------------
# 7. COMPOSITE SCORE + RANKING
# ---------------------------------------------------------------------------
def compute_profile_A_score(df: pd.DataFrame, top_n_per_category: int) -> pd.DataFrame:
    """Combine concept scores into Profile_A_Score and rank within category and overall."""
    weights: dict = {
        "Performance_LT_Score": 0.25,
        "Sharpe_LT_Score": 0.20,
        "Volatility_Score": 0.20,
        "Tracking_Score": 0.10,
        "Cost_Score": 0.10,
        "Quality_Valuation_Score": 0.10,
        "Manager_Score": 0.05,
    }

    df["Profile_A_Score"] = 0.0
    for col, w in weights.items():
        if col in df.columns:
            df["Profile_A_Score"] += w * df[col].fillna(0)

    # Rank within each Morningstar Category (existing behavior)
    df["Profile_A_Rank_In_Category"] = (
        df.groupby("Morningstar Category")["Profile_A_Score"]
        .rank(ascending=False, method="dense")
    )
    df["Profile_A_Selected_Flag"] = df["Profile_A_Rank_In_Category"] <= top_n_per_category

    # NEW: Rank across the entire universe, ignoring category
    df["Profile_A_Rank_Overall"] = df["Profile_A_Score"].rank(ascending=False, method="dense")
    df["Profile_A_Selected_Overall_Flag"] = df["Profile_A_Rank_Overall"] <= top_n_per_category

    return df



# ---------------------------------------------------------------------------
# 8. MAIN
# ---------------------------------------------------------------------------
def main() -> None:
    struct_path, perf_path, out_path, top_n = get_paths_from_user()

    final_out_path: str = build_timestamped_output_path(out_path)
    print(f"Output will be saved to: {final_out_path}")

    print("\nLoading structural data...")
    df_struct: pd.DataFrame = load_structural_data(struct_path, exclude_dir=out_path)

    print("Loading performance data...")
    df_perf: pd.DataFrame = load_performance_data(perf_path, exclude_dir=out_path)

    print("Merging datasets on Ticker...")
    df: pd.DataFrame = merge_datasets(df_struct, df_perf)

    print("Filtering to ETFs only (excluding stocks)...")
    df = apply_etf_only_filter(df)

    print("Building concept scores...")
    df = build_concept_scores(df)

    print("Applying Profile A eligibility filters...")
    df_eligible: pd.DataFrame = apply_profile_A_filters(df)

    print("Computing Profile A composite score and rankings...")
    df_ranked: pd.DataFrame = compute_profile_A_score(df_eligible, top_n)

    # --- TEMPORARY DEBUG: confirm rank-1 count matches category count ---
    category_counts: pd.Series = df_ranked.groupby("Morningstar Category")["Profile_A_Rank_In_Category"].max()
    print(f"\n[DEBUG] Number of distinct categories: {len(category_counts)}")
    print(f"[DEBUG] Category size distribution (max rank per category):")
    print(category_counts.value_counts().sort_index().to_string())
    # --- END DEBUG ---

    print("Saving results...")
    write_excel_with_retry(df_ranked, final_out_path)

    print("Formatting header row...")
    apply_header_formatting(final_out_path)

    display_cols: List[str] = [
        "Ticker", "Name", "Morningstar Category",
        "Profile_A_Score",
        "Profile_A_Rank_In_Category", "Profile_A_Selected_Flag",
        "Profile_A_Rank_Overall", "Profile_A_Selected_Overall_Flag",
    ]
    display_cols = [c for c in display_cols if c in df_ranked.columns]

    selected_by_category: pd.DataFrame = df_ranked[df_ranked["Profile_A_Selected_Flag"]].sort_values(
        ["Morningstar Category", "Profile_A_Rank_In_Category"]
    )
    print(f"\n{len(selected_by_category)} ETFs selected across all categories "
          f"(top {top_n} per category).")
    print(selected_by_category[display_cols].head(50).to_string(index=False))

    selected_overall: pd.DataFrame = df_ranked[df_ranked["Profile_A_Selected_Overall_Flag"]].sort_values(
        "Profile_A_Rank_Overall"
    )
    print(f"\nTop {top_n} ETFs overall (regardless of category):")
    print(selected_overall[display_cols].head(top_n).to_string(index=False))


if __name__ == "__main__":
    main()
