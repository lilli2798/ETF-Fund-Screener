import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

INDEXES = ["^IXIC", "^DJI", "^GSPC"]


def analyze_yearly_returns(csv_file_path, output_file_path=None, db_path=None):
    """
    Analyze yearly ETF returns against index benchmarks.
    
    Args:
        csv_file_path: Path to etf_yearly_history.csv
        output_file_path: Path to save the output Excel file (optional, defaults to csv_path with _analysis suffix and timestamp)
        db_path: Path to SQLite database file (optional, defaults to sources/etf_yearly_returns.db)
    """
    # Generate default output path if not provided
    if output_file_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file_path = csv_file_path.replace('.csv', f'_analysis_{timestamp}.xlsx')
    
    # Read the CSV file
    df = pd.read_csv(csv_file_path, index_col=0)
    
    # Filter to only years (numeric columns)
    year_columns = [col for col in df.columns if col.isdigit()]
    df_years = df[year_columns]
    
    # Get index data
    index_df = df_years.loc[INDEXES]
    
    # Calculate max and min for each year across indexes
    max_per_year = index_df.max(axis=0)
    min_per_year = index_df.min(axis=0)
    
    # Remove indexes from the main dataframe for analysis
    etf_df = df_years.drop(INDEXES, errors='ignore')
    
    # Initialize result dataframe with original data
    result_df = etf_df.copy()
    
    # Calculate metrics
    result_df['total_columns'] = len(year_columns)
    result_df['no_null_columns'] = etf_df.count(axis=1)
    
    # Count better than max and worse than min for each ETF
    better_count = (etf_df > max_per_year).sum(axis=1)
    worst_count = (etf_df < min_per_year).sum(axis=1)
    
    result_df['better'] = better_count
    result_df['worst'] = worst_count
    result_df['better_%'] = result_df['better'] / result_df['no_null_columns']
    result_df['worst_%'] = result_df['worst'] / result_df['no_null_columns']
    result_df['better_worst_diff'] = result_df['better_%'] - result_df['worst_%']
    
    # Generate yearly rankings
    for year in year_columns:
        rank_col = f'Yearly_{year}_rank'
        result_df[rank_col] = etf_df[year].rank(ascending=False, method='dense')
    
    # Calculate percentile buckets
    data_size = len(result_df)
    size_range = data_size / 10
    
    ranking_columns = [f'Yearly_{year}_rank' for year in year_columns]
    ranking_data = result_df[ranking_columns]
    
    # Calculate percentile buckets per year
    percentiles = ['10%', '20%', '30%', '40%', '50%', '60%', '70%', '80%', '90%', 'worse_10%']
    
    # Initialize percentile columns with zeros
    for percentile in percentiles:
        result_df[percentile] = 0
    
    # Calculate percentiles for each year separately
    for year in year_columns:
        rank_col = f'Yearly_{year}_rank'
        
        # Get non-null ranks for this year
        year_ranks = ranking_data[rank_col].dropna()
        
        if len(year_ranks) == 0:
            continue
        
        year_size = len(year_ranks)
        year_size_range = year_size / 10
        
        # Calculate percentile counts for this year
        for i, percentile in enumerate(percentiles):
            if percentile == 'worse_10%':
                # Bottom 10% (ranks from 90% to 100%)
                lower_bound = int(year_size_range * 9) + 1
                upper_bound = int(year_size_range * 10) + 1
            else:
                # Top 10%, 20%, etc. (ranks start at 1 for best)
                # For 10% (i=0): ranks 1 to int(year_size_range * 1)
                # For 20% (i=1): ranks int(year_size_range * 1) + 1 to int(year_size_range * 2)
                lower_bound = int(year_size_range * i) + 1 if i > 0 else 1
                upper_bound = int(year_size_range * (i + 1))
            
            # Count ETFs in this percentile bucket for this year
            in_percentile = ((year_ranks >= lower_bound) & (year_ranks <= upper_bound))
            
            # Add to the corresponding ETFs' percentile counts
            for ticker in year_ranks[in_percentile].index:
                result_df.loc[ticker, percentile] += 1
    
    # Reorder columns as specified
    metric_columns = [
        'total_columns', 'no_null_columns', 'better', 'worst',
        'better_%', 'worst_%', 'better_worst_diff'
    ]
    rank_columns = [f'Yearly_{year}_rank' for year in year_columns]
    percentile_columns = ['10%', '20%', '30%', '40%', '50%', '60%', '70%', '80%', '90%', 'worse_10%']
    
    # Final column order
    final_columns = metric_columns + rank_columns + percentile_columns
    result_df = result_df[final_columns]
    
    # Save to Excel with formatting
    save_with_formatting(result_df, etf_df, max_per_year, min_per_year, output_file_path, year_columns)
    
    return result_df


def save_with_formatting(result_df, etf_df, max_per_year, min_per_year, output_file_path, year_columns):
    """
    Save dataframe to Excel with conditional formatting.
    
    Args:
        result_df: The analysis results dataframe
        etf_df: Original ETF returns dataframe
        max_per_year: Series with max index returns per year
        min_per_year: Series with min index returns per year
        output_file_path: Path to save the Excel file
        year_columns: List of year column names
    """
    # Create Excel file with both sheets (CSV cannot have multiple sheets)
    with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
        etf_df.to_excel(writer, sheet_name='Sheet1', index=True)
        result_df.to_excel(writer, sheet_name='analysis_result', index=True)
    
    # Apply formatting to Sheet1 (raw data)
    wb = load_workbook(output_file_path)
    ws = wb['Sheet1']
    
    # Define fills and fonts
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    gray_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    large_font = Font(size=15)  # 150% larger than default 10pt
    
    # Apply larger font to all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.font = large_font
    
    # Get year column indices (columns B onwards, since A is ticker)
    # Map year to column letter
    header_row = 1
    year_col_map = {}
    for col_idx, cell in enumerate(ws[header_row], 1):
        if cell.value and cell.value in year_columns:
            year_col_map[cell.value] = col_idx
    
    # Iterate through each row (starting from row 2, after header)
    for row_idx in range(2, ws.max_row + 1):
        ticker_cell = ws.cell(row=row_idx, column=1)
        ticker = ticker_cell.value
        
        if ticker and ticker in etf_df.index:
            for year in year_columns:
                if year in year_col_map:
                    col_idx = year_col_map[year]
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell_value = cell.value
                    
                    if pd.notna(cell_value):
                        # Better than max (yellow)
                        if cell_value > max_per_year[year]:
                            cell.fill = yellow_fill
                        # Worse than min (gray)
                        elif cell_value < min_per_year[year]:
                            cell.fill = gray_fill
    
    # Apply larger font to analysis_result sheet as well
    ws_analysis = wb['analysis_result']
    for row in ws_analysis.iter_rows():
        for cell in row:
            cell.font = large_font
    
    wb.save(output_file_path)
    print(f"Analysis saved to {output_file_path} with formatting")


if __name__ == "__main__":
    from config import DEFAULT_CONFIG
    
    csv_path = "/Users/lihongfeng/Library/CloudStorage/OneDrive-YaleUniversity/Projects/ETF-Fund-Screener/etf_screen_yahoo_by_date/yearly_executions/sources/etf_yearly_history.csv"
    
    # Get database path from config
    cache_config = DEFAULT_CONFIG.get("caching", {})
    db_path = cache_config.get("yearly_returns_db_path", None)
    
    result = analyze_yearly_returns(csv_path, db_path=db_path)
    print("Analysis complete!")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    print(f"Results saved to {csv_path.replace('.csv', f'_analysis_{timestamp}.xlsx')}")
    print(result.head())
